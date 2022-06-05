#!/usr/bin/python3 

# Author: Molly Johnson
# Date: 3/24/21
# Description: Creates cboc signin excel file for
# checking in clinics for every month in one year. Will adjust 
# for the days of the month, weekends, and federal holidays

# function comment template
####################################################################
### Function Title:
### Arguments:
### Returns:
### Description: 
###################################################################

# import necessary libraries 
from openpyxl import Workbook
from datetime import datetime
import calendar
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import os
import federalHolidayCalculator
import CBOC
import readFile

# create "constants"
FZ = "Frozen"
TCH = "Tech"
TOA = "Time of Arrival"
SMP = "SMP Check"
CBOC = "_cboc_signin_sheets"
USER_INPUT_LENGTH = 4
MIN_YEAR = 2021
MID_DATE = 15
START_MONTH = 1
END_MONTH = 12
CBOC_COL_WIDTH = 12.5
WEEKEND_AND_HOLIDAY_COL_WIDTH = 2.5
HEADER_ROW_HEIGHT = 40
CBOC_ROW_HEIGHT = 18
DATE_ROW_HEIGHT = 18
TECH_TOA_ROW_HEIGHT = 27
CBOC_NAME_AND_FROZEN_ROW_HEIGHT = 18 
SPACER_ROW_HEIGHT = 6
SMP_ROW_HEIGHT = 18
HEADER_ROW = 1
HEADER_AND_LABELS_COL = 1
CBOC_COL = 1
CBOC_ROW = 2
DATE_ROW = 3
TECH_TOA_ROW = 4
CBOC_NAME_AND_FROZEN_ROW_START = 5

# set "constant" cell border values
THIN = Side(border_style = "thin", color = "000000")
DOUBLE = Side(border_style = "medium", color = "000000")
THICK = Side(border_style = "thick", color = "001C54")
CBOC_NAME_BORDER = Border(top = THICK , left = THICK, right = THICK, bottom = THICK) 
TIMES_NEW_ROMAN_FONT = Font(name = 'Times New Roman', size = 10, bold = True)
DATE_BORDER = Border(left = THICK, right = THICK, bottom = THICK)
BIG_SPACE_BORDER = Border(left = THICK, right = THICK)
SPACER_BORDER = Border(left = THICK, right = THICK)
CBOC_NAME_ONLY_BORDER = Border(top = DOUBLE, left = THICK, right = THICK, bottom = THIN)
FROZEN_ONLY_BORDER = Border(left = THICK, right = THICK, bottom = DOUBLE)
BOTTOM_ROW_BORDER_CBOC_NAME = Border(left = THICK, right = THICK, bottom = THICK)
DATE_FONT = Font(name = 'Calibri', size = 11, bold = True)
TECH_FONT = Font(name = 'Calibri', size = 9)
TOA_FONT = Font(name = 'Calibri', size = 5)
DATE_BORDER_LEFT = Border(top = THICK, left = THICK, bottom = THICK)
DATE_BORDER_RIGHT = Border(top = THICK, right = THICK, bottom = THICK)
TECH_INFO_BORDER_LEFT = Border(left = THICK, right = THIN, bottom = DOUBLE)
TECH_INFO_BORDER_RIGHT = Border(right = THICK, bottom = DOUBLE)
SIGNATURE_BORDER_TOP_LEFT = Border(top = DOUBLE, left = THICK, right = THIN, bottom = THIN)
SIGNATURE_BORDER_BOTTOM_LEFT = Border(left = THICK, right = THIN, bottom = DOUBLE)
SIGNATURE_BORDER_TOP_RIGHT = Border(top = DOUBLE, right = THICK, bottom = THIN)
SIGNATURE_BORDER_BOTTOM_RIGHT = Border(right = THICK, left = THIN, bottom = DOUBLE)
SPACER_BORDER_LEFT = Border(right = THIN, left = THICK)
SPACER_BORDER_RIGHT = Border(right = THICK, left = THIN)
WEEKEND_AND_HOLIDAY_FILL_COLOR = PatternFill(fill_type = "solid", start_color = "BFBFBF", end_color = "BFBFBF")

####################################################################
### Function Title: isValidUserInput()
### Arguments: user input (a 4-digit year in string format)
### Returns: boolean
### Description: checks if the user entered a 4-digit int, and if it's
### a year greater than or equal to the current year (2021). returns
### true if valid, false otherwise
####################################################################
def isValidUserInput(userInput):
    # check that length of user input string is correct
    if(len(userInput) != USER_INPUT_LENGTH):
        return False
    # is / or -, and last 2 chars are digits.
    i = 0
    while(i < USER_INPUT_LENGTH):
        if(userInput[i].isdigit() == False):
        	return False
        i += 1
    # check that year is no earlier than 2021 (year this program written)
    if (int(userInput) < MIN_YEAR):
        return False
    # otherwise met all requirements, return true
    return True

####################################################################
### Function Title: getStartDate()
### Arguments: none
### Returns: start date (in format mm-dd-yyyy) and the year input by
### user (in format yyyy)
### Description: gets input from the user. checks if is a valid year
### in formay yyyy and if is greater than or equal to 2021 (current year).
### if input is invalid, prints error message to the user and keeps asking
### for input until valid input is received. will then use the user's input
### year to create a valid start date that can be used with strptime to get
### a datetime object (which must be in format mm-dd-yyyy). always starts with
### jan 1 of whatever year requested by the user since will create an entire
### year's worth of cboc sheets starting in january
####################################################################
def getStartDate():
    # get start date of the month from user
    while(True):
        userInputYear = input("\nEnter year in the format yyyy: ")
        if(isValidUserInput(userInputYear)):
            break
        else:
            print("Your entry was invalid or a previous year. Enter current year in the format yyyy:")

    # reformat start date input into string for datetime in format 01-01-20yy
    startDate = "01-01-" + userInputYear
    return startDate, userInputYear####################################################################
### Function Title: saveExcelFile()
### Arguments: current month, current year, the workbook, the date
### time object
### Returns: none 
### Description: creates new folder if none exists. changes to that
### directory either way. saves the current workbook to an excel file.
###if the month is a single digit (1 - 9), will add a 0 in front of
### it so the files alphebatize in the correct order in the folder.
####################################################################

def saveExcelFile(currMonth, currYear, wb, dateTimeObj):
    # if is jan, create folder for all the sheets. otherwise you're already in the folder
    if(currMonth == 1):
        # if directory doesn't exist already, create it. then change to that directory from base directory either way
        if(os.path.isdir(str(currYear) + CBOC) == False):
            os.mkdir(str(currYear) + CBOC)
        os.chdir(str(currYear) + CBOC)

    # save workbook to excel file
    if(currMonth < 10):
        strMonth = "0" + str(currMonth)
    else:
        strMonth = str(currMonth)
    wb.save(strMonth + calendar.month_name[dateTimeObj.month] + "_" + str(dateTimeObj.year) + CBOC + '.xlsx')  

####################################################################
### Function Title: main()
### Arguments: none
### Returns: none
### Description: has loop for each month of the given year chosen by
### the user. creates workbook and 2 worksheets per month. calls other
### functions used to get user info, get the datetime object, calculate
### the num of days in the month, calculate the federal holidays, set
### all row and column info for all cells in the cboc signin sheet,
### creates a folder for each year's worksheets (unless the folder
### already exists), and saves all excel spreadsheets using the month
### number/name formatted so that it will have them in correct month
### order in the folder (jan - dec for the given year).
###################################################################
def main():
    # create lists for both SMP and non SMP CBOC names
    smpCBOCs = []
    noSMPCBOCs = []

    # get CBOCs from file
    smpCBOCs, noSMPCBOCs = readFile.getCBOClists(smpCBOCs, noSMPCBOCs)
    
    currMonth = START_MONTH
    currDate, currYear = getStartDate()
    print("...excel spreadsheet creation in progress please wait...")
    
    while (currMonth <= END_MONTH):
        # create workbook (1st sheet at pos 0 created automatically)
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "1-15"
    
        # create 2nd sheet at pos 1
        ws2 = wb.create_sheet("16-End", 1)
        
        # Create date object in format mm-dd-yyyy from start date string
        dateTimeObj = datetime.strptime(currDate, "%m-%d-%Y")
        
        # get end date for the month
        endDate = calendar.monthrange(dateTimeObj.year, dateTimeObj.month)[1]
        
        # calculate holiday dates for the month
        holidayDates = []
        holidayDates = federalHolidayCalculator.calcFedHolidays(dateTimeObj) 


        #####################

        #####################
        # save the current month's excel file
        saveExcelFile(currMonth, currYear, wb, dateTimeObj)
        
        #increment the month
        currMonth += 1
        currDate = str(currMonth) + "-01-" + str(currYear)
    
    # tell user what folder their files are in. wait for their input (enter) before exiting.
    print("\nSpreadsheets created, see folder: \"" + str(currYear) + "_cboc_signin_sheets\" for your files.")
    print("Press \"Enter\" to finish.")
    input()

if __name__ == "__main__":
    main()